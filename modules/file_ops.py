# -*- coding: utf-8 -*-
"""
到期提醒工具 — 文件加载 / 写回 / 导出 / 送货 Mixin
"""

import os
import sys
import shutil
import threading
import traceback
from collections import Counter
from datetime import datetime

from tkinter import messagebox, filedialog, simpledialog

import openpyxl
try:
    import xlrd
except ImportError:
    xlrd = None

from modules.config import COL_MAP
from modules.utils import _row_value, detect_date_columns, detect_col_map, log_debug, log_error, log_trace, parse_date, clear_date_cache


class FileOpsMixin:
    """文件浏览、加载、写回 Excel、导出、送货。"""

    # -----------------------------------------------------------
    # 文件
    # -----------------------------------------------------------

    def _open_file(self):
        log_trace("[FILE] _open_file 调用")
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("所有 Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm *.xls *.xlt"),
                ("新版 Excel (.xlsx/.xlsm/.xltx/.xltm)", "*.xlsx *.xlsm *.xltx *.xltm"),
                ("老版 Excel (.xls/.xlt)", "*.xls *.xlt"),
                ("所有文件", "*.*"),
            ],
        )
        if path:
            log_trace(f"[FILE] _open_file 选择路径: {path}")
            self.path_var.set(path)
            self._load_workbook(path)

    def _reload_current_file(self):
        p = self._loaded_path or self.path_var.get()
        log_trace(f"[FILE] _reload_current_file: path={p}, loading={self._loading}, checking={self._checking}")
        if not p:
            self.status_var.set("当前没有加载文件可重新加载"); return
        if not os.path.exists(p):
            messagebox.showwarning("重新加载失败", f"文件不存在：\n{p}"); return
        if self._loading or self._checking:
            messagebox.showinfo("提示", "正在处理中,请稍候"); return
        self.status_var.set(f"正在重新加载：{os.path.basename(p)}")
        self._load_workbook(p, keep_date_col=True, run_analysis=True)

    def _open_file_folder(self):
        p = self._loaded_path or self.path_var.get()
        if not p:
            messagebox.showwarning("提示", "请先加载一个文件"); return
        folder = os.path.dirname(p) if os.path.exists(p) else p
        folder = folder if os.path.exists(folder) else os.path.expanduser("~")
        self._shell_open(folder, "打开文件夹失败")

    def _open_in_excel(self):
        p = self._loaded_path or self.path_var.get()
        if not p:
            messagebox.showwarning("提示", "请先加载一个文件"); return
        if not os.path.exists(p):
            messagebox.showwarning("打开失败", f"文件不存在：\n{p}"); return
        self._shell_open(p, "打开失败")
        self.status_var.set(f"已用系统默认程序打开：{os.path.basename(p)}")

    def _shell_open(self, path, err_title):
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)
            elif sys.platform == "darwin":
                os.system(f'open "{path}"')
            else:
                os.system(f'xdg-open "{path}"')
        except Exception as e:
            messagebox.showerror(err_title, str(e))

    def _focus_history_search(self):
        if hasattr(self, "notebook") and hasattr(self, "history_tab"):
            self.notebook.select(self.history_tab)
        if hasattr(self, "history_search_var"):
            self.history_search_var.set("")
            try:
                self.history_search_entry.focus_set()
            except Exception:
                pass

    def _load_last_file(self):
        p = self.path_var.get()
        log_trace(f"[FILE] _load_last_file: path={p}, exists={os.path.exists(p) if p else False}")
        if p and os.path.exists(p):
            self._load_workbook(p)
        elif p:
            self.path_var.set("")
            self.status_var.set(f"上次文件已失效,已清理: {os.path.basename(p)}")

    # -----------------------------------------------------------
    # 加载 Excel
    # -----------------------------------------------------------

    def _load_workbook(self, path, keep_date_col=False, run_analysis=False):
        log_trace(f"[LOAD] _load_workbook 开始: path={path}, _loading={self._loading}, keep_date_col={keep_date_col}, run_analysis={run_analysis}")
        if self._loading:
            log_trace("[LOAD] _load_workbook 跳过: _loading=True")
            return
        # ── 保存当前日期列的列号标识（1-based，跨加载稳定） ──
        saved_date_col_key = None
        if keep_date_col and self._date_cols:
            idx = self.date_col_cb.current()
            if 0 <= idx < len(self._date_cols):
                saved_date_col_key = self._date_cols[idx]["col"]
                log_debug(f"[LOAD] 保存日期列: idx={idx}, col={saved_date_col_key}")
        self._loading = True
        self._sync_button_states()
        self.hint_var.set("正在加载 Excel 文件…")
        self.status_var.set(f"正在读取：{os.path.basename(path)}")

        def worker():
            wb = None
            try:
                clear_date_cache()
                ext = os.path.splitext(path)[1].lower()
                if ext in (".xls", ".xlt"):
                    if xlrd is None:
                        raise ImportError("缺少 xlrd 库，无法读取 .xls 文件。请安装：pip install xlrd")
                    wb = xlrd.open_workbook(path)
                    ws = wb.sheet_by_index(0)
                    headers = tuple(ws.cell_value(0, c) for c in range(ws.ncols))
                    data_rows = []
                    for row_num in range(1, ws.nrows):
                        row = tuple(ws.cell_value(row_num, c) for c in range(ws.ncols))
                        if any(v not in (None, "", 0) for v in row):
                            data_rows.append({"excel_row": row_num + 1, "values": row})
                else:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    ws = wb[wb.sheetnames[0]]
                    row_iter = ws.iter_rows(values_only=True)
                    headers = tuple(next(row_iter, ()))
                    data_rows = []
                    for row_num, row in enumerate(row_iter, start=2):
                        row = tuple(row)
                        if any(v not in (None, "") for v in row):
                            data_rows.append({"excel_row": row_num, "values": row})
                log_debug(f"[LOAD] 读取完成: headers={headers}")
                date_cols = detect_date_columns(headers, [r["values"] for r in data_rows])
                log_debug(f"[LOAD] 数据行数: {len(data_rows)}, 日期列: {len(date_cols)} 个")
                items = [f"{openpyxl.utils.get_column_letter(c['col']) if ext not in ('.xls', '.xlt') else '列'+str(c['col'])}  {c['header']}  ({c['ok']}/{c['total']})" for c in date_cols]

                def update():
                    self.path_var.set(path)
                    self._loaded_path = path
                    self._sheet_headers = headers
                    self._sheet_rows = data_rows
                    self._col_map = detect_col_map(headers)
                    log_debug(f"[LOAD] update: _loaded_path={self._loaded_path}, rows={len(data_rows)}, col_map={self._col_map}")
                    self._result_rows = []
                    self._result_detail_rows = []
                    self._failed_rows = []
                    self._skip_info = {}
                    self._shipped_rows = []
                    self._cached_full_tree_data = None  # 清除树缓存
                    self._date_cols = date_cols
                    self.date_col_cb["values"] = items
                    if items:
                        restored = False
                        if saved_date_col_key is not None:
                            for i, dc in enumerate(date_cols):
                                if dc["col"] == saved_date_col_key:
                                    self._suppress_date_col_change = True
                                    self.date_col_cb.current(i)
                                    self._suppress_date_col_change = False
                                    self.hint_var.set(f"✅ 自动检测 {len(items)} 个日期列，已恢复选择")
                                    log_debug(f"[LOAD] 已恢复日期列: col={saved_date_col_key} → idx={i}")
                                    restored = True
                                    break
                        if not restored:
                            self._suppress_date_col_change = True
                            self.date_col_cb.current(0)
                            self._suppress_date_col_change = False
                            self.hint_var.set(f"✅ 自动检测 {len(items)} 个日期列，已选第 1 个")
                    else:
                        self.date_col_cb.set("")
                        self.hint_var.set("⚠️ 未检测到日期列")
                    self.compare_var.set("历史对比：暂无")
                    self.tv.delete(*self.tv.get_children())
                    self.status_var.set(f"已加载：{os.path.basename(path)}  共 {len(data_rows)} 条数据")
                    self._loading = False
                    self._sync_button_states()
                    self._refresh_shipping_tab()
                    log_debug("[LOAD] update 完成")
                    if run_analysis:
                        log_debug("[LOAD] 调用 _run_analysis")
                        self._run_analysis(show_dialog=False)
                self.after(0, update)
            except Exception as e:
                log_error(f"[LOAD] _load_workbook 异常: {e}")
                log_error(traceback.format_exc())
                def show_error():
                    self._loaded_path = ""
                    self._sheet_headers = ()
                    self._sheet_rows = []
                    self._result_rows = []
                    self._result_detail_rows = []
                    self._failed_rows = []
                    self._skip_info = {}
                    self._shipped_rows = []
                    self._date_cols = []
                    self._col_map = dict(COL_MAP)
                    self.date_col_cb["values"] = []
                    self.date_col_cb.set("")
                    self._loading = False
                    self.hint_var.set("文件加载失败")
                    self.status_var.set("打开文件失败")
                    self._sync_button_states()
                    msg = str(e)
                    if "locked" in msg.lower() or "permission" in msg.lower():
                        msg = "文件已被其他程序打开，请先关闭后再试"
                    elif "invalid" in msg.lower() or "corrupt" in msg.lower():
                        msg = "文件格式损坏，不是有效的 Excel 文件"
                    elif "not a zip" in msg.lower():
                        msg = "文件不是有效的 Excel 格式（.xlsx/.xlsm），请确认文件完整"
                    self._show_analysis_failure(f"无法打开文件：{msg}")
                self.after(0, show_error)
            finally:
                if wb is not None:
                    try: wb.close()
                    except Exception: pass
        threading.Thread(target=worker, daemon=True).start()

    # -----------------------------------------------------------
    # 内存状态同步
    # -----------------------------------------------------------

    @staticmethod
    def _set_tuple_value(t, idx, value):
        lst = list(t)
        if idx < len(lst):
            lst[idx] = value
        else:
            while len(lst) <= idx:
                lst.append("")
            lst[idx] = value
        return tuple(lst)

    def _sync_memory_from_file(self, path):
        """从已保存的文件重新读取以同步内存数据（批量读取优化）。"""
        log_debug(f"[SYNC] _sync_memory_from_file 开始: path={path}")
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            all_rows = list(ws.iter_rows(values_only=True))
            log_debug(f"[SYNC] 文件行数={len(all_rows)}, _sheet_rows数={len(self._sheet_rows)}")
            col_status = self._col_map.get("order_status", 4)
            for idx, row_rec in enumerate(self._sheet_rows):
                row_num = row_rec["excel_row"]
                if row_num <= len(all_rows):
                    values = tuple(all_rows[row_num - 1])
                else:
                    values = tuple(ws.cell(row=row_num, column=c).value
                                   for c in range(1, ws.max_column + 1))
                old_status = row_rec["values"][col_status] if col_status < len(row_rec["values"]) else "?"
                new_status = values[col_status] if col_status < len(values) else "?"
                if old_status != new_status:
                    log_debug(f"[SYNC] row {row_num} 状态变更: {old_status!r} → {new_status!r}")
                self._sheet_rows[idx] = {"excel_row": row_num, "values": values}
            self._shipping_cache_key = None
            wb.close()
            log_debug("[SYNC] _sync_memory_from_file 完成")
        except Exception as e:
            log_error(f"[SYNC] _sync_memory_from_file 失败: {e}")
            log_error(traceback.format_exc())

    # -----------------------------------------------------------
    # 写回 Excel（支持 .xls 自动转 .xlsx）
    # -----------------------------------------------------------

    def _save_new_xlsx_for_xls(self, path, apply_func):
        """将 .xls 文件整体转存为 .xlsx，apply_func(ws) 修改单元格。返回 (新路径, backup_msg)。"""
        base, _ = os.path.splitext(path)
        new_path = base + ".xlsx"
        if os.path.exists(new_path):
            shutil.copy2(new_path, new_path + ".bak")
        wb_old = xlrd.open_workbook(path)
        wb_new = openpyxl.Workbook()
        try:
            sh = wb_old.sheet_by_index(0)
            ws_new = wb_new.active
            for r in range(sh.nrows):
                for c in range(sh.ncols):
                    cell_type = sh.cell_type(r, c)
                    if cell_type in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                        continue
                    cell_value = sh.cell_value(r, c)
                    if cell_type == xlrd.XL_CELL_DATE:
                        date_tuple = xlrd.xldate_as_tuple(cell_value, wb_old.datemode)
                        ws_new.cell(row=r + 1, column=c + 1).value = datetime(*date_tuple)
                    else:
                        ws_new.cell(row=r + 1, column=c + 1).value = cell_value
            apply_func(ws_new)
            wb_new.save(new_path)
        finally:
            try: wb_old.release_resources()
            except Exception: pass
            try: wb_new.close()
            except Exception: pass
        return new_path, f"(已转为 .xlsx 格式保存)"

    def _write_to_workbook(self, path, apply_func, success_message, reload_if_current=False):
        """通用写回方法：apply_func(ws) 负责对 worksheet 执行修改。"""
        log_debug(f"[WRITE] _write_to_workbook 开始: path={path}, reload_if_current={reload_if_current}")
        log_debug(f"[WRITE] _loaded_path={self._loaded_path}, _writing={self._writing}")
        if self._writing:
            messagebox.showinfo("提示", "正在写回原始文件，请稍候"); return
        self._writing = True
        self._sync_button_states()
        self.status_var.set("正在同步写回原始文件…")

        def worker():
            wb = None
            try:
                log_debug("[WRITE] worker 开始: 备份文件...")
                backup_msg = self._backup_file(path)
                is_xls = path.lower().endswith(".xls") and not path.lower().endswith(".xlsx")
                if is_xls:
                    backup_msg = "(原 .xls 文件保留)"
                new_path = path

                if is_xls:
                    new_path, extra_msg = self._save_new_xlsx_for_xls(path, apply_func)
                    backup_msg += " " + extra_msg
                else:
                    log_debug(f"[WRITE] 打开工作簿: {path}")
                    wb = openpyxl.load_workbook(path)
                    ws = wb[wb.sheetnames[0]]
                    apply_func(ws)
                    wb.save(path)
                    log_debug(f"[WRITE] 保存成功: {path}")

                def done():
                    log_debug(f"[WRITE] done() 开始: reload_if_current={reload_if_current}, new_path={new_path}")
                    self._writing = False
                    self._sync_button_states()
                    self.status_var.set(success_message)
                    if reload_if_current:
                        # ── 保存 Treeview 选中状态（日期列由 _load_workbook 内部恢复） ──
                        saved_row_nums = set()
                        for iid in self.tv.selection():
                            payload = self._tree_item_map.get(iid)
                            if payload and payload.get("kind") == "detail":
                                saved_row_nums.add(payload["row"]["row_num"])
                        if saved_row_nums:
                            self._pending_selection_restore = saved_row_nums
                            log_debug(f"[WRITE] 保存 Treeview 选中: {saved_row_nums}")
                        log_debug(f"[WRITE] done() 重新加载文件: {new_path}")
                        self._loaded_path = new_path
                        self.path_var.set(new_path)
                        self._load_workbook(new_path, keep_date_col=True, run_analysis=True)
                    messagebox.showinfo("完成", f"{success_message}\n{backup_msg}")
                    log_debug("[WRITE] done() 完成")
                self.after(0, done)
            except PermissionError:
                log_error("_write_to_workbook: PermissionError")
                self.after(0, lambda: self._write_error("文件已被其他程序（如 Excel）打开，请先关闭后再试"))
            except Exception as e:
                log_error(f"_write_to_workbook 失败: {e}")
                log_error(traceback.format_exc())
                def show_error():
                    self._writing = False; self._sync_button_states()
                    self.status_var.set("写回原始文件失败")
                    msg = str(e)
                    if "locked" in msg.lower() or "permission" in msg.lower():
                        msg = "文件已被其他程序打开，请先关闭后再试"
                    messagebox.showerror("写回失败", f"无法写入文件：\n{msg}")
                self.after(0, show_error)
            finally:
                if wb is not None:
                    try: wb.close()
                    except Exception: pass
        threading.Thread(target=worker, daemon=True).start()

    def _write_status_updates(self, path, updates, success_message, reload_if_current=False, col_idx=None):
        """写入订单状态更新，列号由 col_idx 决定（默认使用 self._col_map）。"""
        if col_idx is None:
            col_idx = self._col_map.get("order_status", 4)
        def apply(ws):
            for row_num, value in updates.items():
                ws.cell(row=row_num, column=col_idx + 1).value = value
        self._write_to_workbook(path, apply, success_message, reload_if_current)

    def _write_error(self, msg):
        self._writing = False
        self._sync_button_states()
        self.status_var.set("写回原始文件失败")
        messagebox.showerror("写回失败", msg)

    def _backup_file(self, path):
        try:
            base, ext = os.path.splitext(path)
            backup_path = f"{base}.bak{ext}"
            shutil.copy2(path, backup_path)
            return f"(已备份到 {os.path.basename(backup_path)})"
        except Exception as e:
            log_error(f"备份失败: {e}")
            return "(备份失败, 继续保存)"

    def _edit_selected_status(self):
        log_debug("[EDIT] _edit_selected_status 调用")
        selected = self.tv.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择一条明细记录"); return
        payload = self._tree_item_map.get(selected[0])
        if not payload or payload.get("kind") != "detail":
            messagebox.showwarning("提示", "请在展开后的明细行上修改订单状态"); return
        detail = payload["row"]
        col_status = self._col_map.get("order_status", 4)
        new_status = simpledialog.askstring(
            "修改订单状态",
            f"第 {detail['row_num']} 行，客户：{detail['customer']}，产品：{detail['product']}，请输入新的订单状态：",
            initialvalue=detail["order_status"],
            parent=self,
        )
        if new_status is None: return
        new_status = new_status.strip()
        if not new_status:
            messagebox.showwarning("提示", "订单状态不能为空"); return
        if new_status == detail["order_status"]: return
        log_debug(f"[EDIT] 修改状态: row={detail['row_num']}, {detail['order_status']!r} → {new_status!r}")
        self._write_status_updates(
            self._loaded_path,
            {detail["row_num"]: new_status},
            f"第 {detail['row_num']} 行订单状态已更新为：{new_status}",
            reload_if_current=True,
        )

    # -----------------------------------------------------------
    # 一键送货
    # -----------------------------------------------------------

    def _ship_now(self):
        log_debug("========== _ship_now 开始 ==========")
        log_debug(f"_loaded_path={self._loaded_path}, _result_detail_rows={len(self._result_detail_rows)}, _writing={self._writing}")
        selected = self.tv.selection()
        log_debug(f"tv.selection() = {selected}")
        if not selected:
            messagebox.showwarning("提示", "请先在分析结果中选择一条明细记录"); return
        details = []
        for iid in selected:
            payload = self._tree_item_map.get(iid)
            if payload and payload.get("kind") == "detail":
                details.append(payload["row"])
        if not details:
            messagebox.showwarning("提示", "请在展开后的明细行上操作（聚合行/产品行不支持发货）"); return
        col_status = self._col_map.get("order_status", 4)
        idx = self.date_col_cb.current()
        if idx < 0 or idx >= len(self._date_cols):
            messagebox.showwarning("提示", "请先选择日期列"); return
        date_col_idx = self._date_cols[idx]["index"]
        date_col_letter = openpyxl.utils.get_column_letter(date_col_idx + 1)
        if not self._loaded_path:
            messagebox.showwarning("提示", "未加载文件,无法写回"); return
        cnt = Counter((d.get("order_status") or "空") for d in details)
        preview = " / ".join(f"{k}{v}" for k, v in cnt.most_common(3))
        today_str = datetime.now().strftime("%Y-%m-%d")
        if not messagebox.askyesno("确认一键送货",
            f"将对 {len(details)} 条记录执行：\n"
            f"  · 订单状态({openpyxl.utils.get_column_letter(col_status + 1)} 列) → 已发货\n"
            f"  · 送货日期({date_col_letter} 列) → {today_str}\n"
            f"原状态分布: {preview}\n"
            f"写回前会自动备份到 .bak 文件,确认继续？"):
            return
        log_debug(f"_ship_now: 准备写回 {len(details)} 条, 列={date_col_idx}, 路径={self._loaded_path}")
        def apply(ws):
            for d in details:
                ws.cell(row=d["row_num"], column=col_status + 1).value = "已发货"
                ws.cell(row=d["row_num"], column=date_col_idx + 1).value = today_str
        self._write_to_workbook(
            self._loaded_path, apply,
            f"已发货 {len(details)} 条,日期已改为今天({today_str})",
            reload_if_current=True,
        )

    # -----------------------------------------------------------
    # 送货记录
    # -----------------------------------------------------------

    def _refresh_shipping_tab(self):
        log_debug(f"[SHIP] _refresh_shipping_tab: _sheet_rows={len(self._sheet_rows)}")
        # 缓存：数据未变更则跳过重建
        cache_key = (id(self._sheet_rows), len(self._sheet_rows))
        if not self._sheet_rows:
            self.shipping_info_var.set("送货记录：暂无数据")
            self._sync_button_states()
            self._shipping_cache_key = None
            return
        if getattr(self, "_shipping_cache_key", None) == cache_key:
            return
        self._shipping_cache_key = cache_key
        for item in self.shipping_tv.get_children():
            self.shipping_tv.delete(item)
        self._shipped_rows = []
        # 使用用户选择的日期列（而非固定第 0 个检测列）
        sel_idx = self.date_col_cb.current()
        date_ci = self._date_cols[sel_idx]["index"] if 0 <= sel_idx < len(self._date_cols) else (
            self._date_cols[0]["index"] if self._date_cols else None)
        for row_rec in self._sheet_rows:
            row = row_rec["values"]
            if "已发货" in str(_row_value(row, self._col_map.get("order_status", 4)) or ""):
                delivery_date = ""
                if date_ci is not None:
                    raw_date = _row_value(row, date_ci)
                    if raw_date is not None:
                        parsed, _ = parse_date(raw_date)
                        if parsed:
                            delivery_date = parsed.strftime("%Y-%m-%d")
                        else:
                            delivery_date = str(raw_date)[:10]
                self._shipped_rows.append((
                    delivery_date,
                    str(_row_value(row, self._col_map.get("customer", 0)) or ""),
                    str(_row_value(row, self._col_map.get("product", 5)) or ""),
                    str(_row_value(row, self._col_map.get("order_status", 4)) or ""),
                    str(_row_value(row, self._col_map.get("note", 10)) or ""),
                ))
        for r in self._shipped_rows:
            self.shipping_tv.insert("", "end", values=r, tags=("shipped",))
        self.shipping_info_var.set(f"送货记录：共 {len(self._shipped_rows)} 条已发货")
        self._sync_button_states()

    def _export_shipping(self):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        base = f"送货记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        fpath = self._unique_path(desktop, base, ".xlsx")
        if not self._shipped_rows:
            messagebox.showwarning("提示", "没有送货数据可导出"); return
        wb = None
        try:
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "送货记录"
            ws.append(["送货日期", "客户", "产品", "订单状态", "备注"])
            for r in self._shipped_rows:
                ws.append(list(r))
            wb.save(fpath)
            self.status_var.set(f"✅ 送货记录已导出到桌面：{os.path.basename(fpath)}")
            messagebox.showinfo("导出成功", f"文件已保存到桌面：{os.path.basename(fpath)}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
        finally:
            if wb is not None:
                try: wb.close()
                except Exception: pass

    # -----------------------------------------------------------
    # 导出分析结果
    # -----------------------------------------------------------

    def _export(self):
        log_debug("[EXPORT] _export 调用")
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        base = f"到期提醒_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        fpath = self._unique_path(desktop, base, ".xlsx")
        rows = list(self._result_rows)
        if not rows:
            messagebox.showwarning("提示", "没有数据可导出"); return
        if self._exporting:
            messagebox.showinfo("提示", "正在导出，请稍候"); return
        self._exporting = True
        self._sync_button_states()
        self.status_var.set("正在导出结果…")
        self.export_btn.config(text="导出中…")
        self.bottom_export_btn.config(text="导出中…")

        def worker():
            wb = None
            try:
                wb = openpyxl.Workbook()
                ws = wb.active; ws.title = "到期提醒"
                ws.append(["状态", "剩余", "日期", "客户", "产品", "订单状态", "备注"])
                for r in rows:
                    ws.append(list(r))
                wb.save(fpath)
                fname = os.path.basename(fpath)

                def done():
                    self._exporting = False
                    self.export_btn.config(text="导出结果")
                    self.bottom_export_btn.config(text="💾  导出到桌面")
                    self._sync_button_states()
                    self.status_var.set(f"✅ 已保存到桌面：{fname}")
                    messagebox.showinfo("导出成功", f"文件已保存到桌面：{fname}")
                self.after(0, done)
            except Exception as e:
                def show_error():
                    self._exporting = False
                    self.export_btn.config(text="导出结果")
                    self.bottom_export_btn.config(text="💾  导出到桌面")
                    self._sync_button_states()
                    self.status_var.set("导出失败")
                    messagebox.showerror("导出失败", str(e))
                self.after(0, show_error)
            finally:
                if wb is not None:
                    try: wb.close()
                    except Exception: pass
        threading.Thread(target=worker, daemon=True).start()

    @staticmethod
    def _unique_path(folder, base, ext):
        fpath = os.path.join(folder, f"{base}{ext}")
        if not os.path.exists(fpath):
            return fpath
        for i in range(1, 1000):
            candidate = os.path.join(folder, f"{base}_{i}{ext}")
            if not os.path.exists(candidate):
                return candidate
        return fpath
